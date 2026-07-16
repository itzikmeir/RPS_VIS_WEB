"""
Build a duplicate of Models_Experiment_Order_Expanded_V2.xlsx with per-route,
per-segment score/time/length columns added to the "קטלוג תרחישים" sheet,
extracted directly from the scenario HTML files.

Each route (A/B/C) has 3 segments; routeScores are verified identical across
the Correct_Scenarios / Inaccurate_Scenarios folders and across the H/R/S
visualization variants of the same scenario, so one representative HTML file
per scenario is enough.

New columns per route (20 per route, 60 total), e.g. for route A:
  A1_RECEPTION_SCORE, A1_ECONOMY_SCORE, A1_SCENIC_SCORE, A1_SPEED_SCORE,
  A1_TIME, A1_LENGTH   (repeated for segments 1-3)
  A_TOT_TIME, A_TOT_LENGTH
TIME is in minutes (rounded to 2 decimals); LENGTH is in meters (source unit).

Usage:
  python python_scripts/build_route_scores_excel.py
"""
import re
import json
import glob
from pathlib import Path

import openpyxl

ROOT_DIR = Path(__file__).resolve().parent.parent
SRC_XLSX = ROOT_DIR / "Model_Ordered_experiment" / "Models_Experiment_Order_Expanded_V2.xlsx"
OUT_XLSX = ROOT_DIR / "Model_Ordered_experiment" / "Models_Experiment_Order_Expanded_V2_with_route_scores.xlsx"
SCENARIOS_DIR = ROOT_DIR / "Scenarios" / "Correct_Scenarios"

CATEGORY_MAP = [
    ("RECEPTION_SCORE", "commScore"),
    ("ECONOMY_SCORE", "economyScore"),
    ("SCENIC_SCORE", "scenicScore"),
    ("SPEED_SCORE", "speedScore"),
]
ROUTES = ["A", "B", "C"]
SEGMENTS = [1, 2, 3]


def extract_data(path: Path):
    content = path.read_text(encoding="utf-8")
    m = re.search(r"const DATA = (\{.*?\});", content, re.DOTALL)
    if not m:
        return None
    return json.loads(m.group(1))


def build_scenario_scores() -> dict:
    """One representative HTML file per base scenario (SCN_XXX)."""
    files = sorted(glob.glob(str(SCENARIOS_DIR / "SCN_*.html")))
    by_scn = {}
    for f in files:
        fname = Path(f).name
        m = re.match(r"(SCN_\d+)_([HRS])\.html", fname)
        if not m:
            continue
        scn_base = m.group(1)
        if scn_base in by_scn:
            continue
        data = extract_data(Path(f))
        if data is None:
            print(f"[WARN] no DATA blob in {f}")
            continue
        by_scn[scn_base] = data
    return by_scn


def scores_for_scenario(data: dict) -> dict:
    route_scores = data.get("routeScores", [])
    by_route_seg = {(row["route"], row["segment"]): row for row in route_scores}

    out = {}
    for route in ROUTES:
        total_time_s = 0.0
        total_len_m = 0.0
        for seg in SEGMENTS:
            row = by_route_seg.get((route, seg))
            if row is None:
                continue
            for col_suffix, field in CATEGORY_MAP:
                out[f"{route}{seg}_{col_suffix}"] = round(row[field], 2)
            out[f"{route}{seg}_TIME"] = round(row["timeS"] / 60.0, 2)
            out[f"{route}{seg}_LENGTH"] = round(row["lengthM"], 1)
            total_time_s += row["timeS"]
            total_len_m += row["lengthM"]
        out[f"{route}_TOT_TIME"] = round(total_time_s / 60.0, 2)
        out[f"{route}_TOT_LENGTH"] = round(total_len_m, 1)
    return out


def ordered_new_columns() -> list:
    cols = []
    for route in ROUTES:
        for seg in SEGMENTS:
            for col_suffix, _ in CATEGORY_MAP:
                cols.append(f"{route}{seg}_{col_suffix}")
            cols.append(f"{route}{seg}_TIME")
            cols.append(f"{route}{seg}_LENGTH")
        cols.append(f"{route}_TOT_TIME")
        cols.append(f"{route}_TOT_LENGTH")
    return cols


def main():
    scenario_data = build_scenario_scores()
    print(f"Extracted routeScores for {len(scenario_data)} scenarios")
    missing = {f"SCN_{i:03d}" for i in range(1, 34)} - set(scenario_data.keys())
    if missing:
        print("[WARN] missing scenarios (no HTML found):", sorted(missing))

    scenario_scores = {scn: scores_for_scenario(d) for scn, d in scenario_data.items()}
    new_cols = ordered_new_columns()
    print(f"New columns to add: {len(new_cols)}")

    # data_only=True: read the last-computed VALUE of formula cells (e.g. Scenario_ID_H
    # is the formula =A2&"_H") rather than the formula text. Loading with
    # data_only=False and saving back would keep the formula but drop its cached
    # value, leaving those cells blank to any reader that doesn't recalculate
    # (pandas, openpyxl, etc.) until the file is next opened in real Excel.
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb["קטלוג תרחישים"]

    start_col = ws.max_column + 1
    for i, col_name in enumerate(new_cols):
        ws.cell(row=1, column=start_col + i, value=col_name)

    matched = 0
    unmatched = []
    for r in range(2, ws.max_row + 1):
        scn_id = ws.cell(r, 1).value
        if not scn_id:
            continue
        scores = scenario_scores.get(scn_id)
        if scores is None:
            unmatched.append(scn_id)
            continue
        for i, col_name in enumerate(new_cols):
            ws.cell(row=r, column=start_col + i, value=scores.get(col_name))
        matched += 1

    print(f"Matched {matched} catalog rows; unmatched: {unmatched}")

    wb.save(OUT_XLSX)
    print(f"Saved: {OUT_XLSX}")


if __name__ == "__main__":
    main()
